#!/usr/bin/env python3
"""
ED Leaderboard weekly updater.

Usage:
    python3 update_leaderboard.py
        -> automatically uses the most recently modified .xlsx anywhere
           under "Provider PPH data" (next to this repo folder)

    python3 update_leaderboard.py /path/to/spreadsheet.xlsx
        -> uses a specific file instead

What it does, every time, in this order:
  1. Reads every weekly sheet in Wendy's spreadsheet (Cumulative tabs and
     tabs without ESI columns are skipped)
  2. Figures out which weeks are NOT already in src/data.js
  3. Builds week objects in the exact SEED_WEEKS schema
  4. Prepends new weeks (newest first) to src/data.js
  5. If any sheet has a "Provider Name" column, syncs DEFAULT_NAMES
     in src/data.js from the spreadsheet (sheet is source of truth)
  6. Bumps DATA_VERSION in src/App.jsx by 1 if anything changed
  7. Prints a summary report with any data anomalies flagged

Sheets that have a Provider Name column but no Letter column are handled
by looking the name up in DEFAULT_NAMES from data.js. Unmatched names are
reported loudly and skipped.

It does NOT commit or push. That stays a human (or supervised-agent) decision.
Safe to run twice: already-present weeks and unchanged names are skipped.
"""

import os
import re
import sys
import datetime

REPO = os.path.dirname(os.path.abspath(__file__))
DATA_JS = os.path.join(REPO, "src", "data.js")
APP_JSX = os.path.join(REPO, "src", "App.jsx")

MONTHS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Anomaly thresholds, tune as needed
MIN_PTHR_FLAG = 1.0     # active week below this pt/hr gets flagged
MIN_HOURS_FLAG = 6      # active week with fewer hours than this gets flagged

# Column aliases. Matched after lowercasing and removing all spaces.
COL_ALIASES = {
    "letter": ["letter"],
    "name":   ["providername", "provider", "name", "physician"],
    "hours":  ["hoursworked", "hours", "hrs"],
    "pts":    ["#pts", "pts", "#ptsw/oapp", "patients", "#patients"],
    "pthr":   ["pt/hr", "pthr", "pts/hr"],
    "esi1":   ["esi1"], "esi2": ["esi2"], "esi3": ["esi3"],
    "esi4":   ["esi4"], "esi5": ["esi5"],
}


def norm(v):
    return str(v).strip().lower().replace(" ", "") if v is not None else ""


def parse_sheet_dates(sheet_name):
    """Parse sheet names like '6.21-6.27.26', '5.17.26-5.23.26',
    '7.26-8.' (end missing day) or '.2-8.8.26' (start missing month)
    into (start_date, end_date). A Sun-Sat week is assumed when one
    half is incomplete."""
    halves = sheet_name.split("-")
    if len(halves) != 2:
        raise ValueError(f"Can't parse sheet name: {sheet_name}")

    def toks(half):
        return [int(t) for t in re.findall(r"\d+", half)]

    t1, t2 = toks(halves[0]), toks(halves[1])
    year = 2026
    for t in (t2, t1):  # a 2-digit year only counts as the 3rd token of m.d.yy
        if len(t) >= 3 and 12 < t[2] < 100:
            year = 2000 + t[2]
            break

    start = end = None
    if len(t1) >= 2:
        start = datetime.date(year, t1[0], t1[1])
    if len(t2) >= 2:
        end = datetime.date(year, t2[0], t2[1])
        if start and end < start:
            end = datetime.date(year + 1, t2[0], t2[1])

    if start and not end:
        end = start + datetime.timedelta(days=6)
    elif end and not start:
        start = end - datetime.timedelta(days=6)
    elif not start and not end:
        raise ValueError(f"Can't parse sheet name: {sheet_name}")
    return start, end


def week_id(start):
    """ID convention: ISO week number of the Monday inside a Sun-Sat week."""
    monday = start + datetime.timedelta(days=1) if start.weekday() == 6 else start
    iso = monday.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}".replace("-W0", "-W") if iso[1] >= 10 else f"{iso[0]}-W{iso[1]}"


def labels_for(start, end):
    label = f"Week of {start.month}/{start.day}\u2013{end.month}/{end.day}"
    if start.month == end.month:
        drange = f"{MONTHS[start.month]} {start.day}\u2013{end.day}, {end.year}"
    else:
        drange = (f"{MONTHS[start.month]} {start.day}\u2013"
                  f"{MONTHS[end.month]} {end.day}, {end.year}")
    return label, drange


def find_header(rows):
    """Return (index, normalized header list) of the first row that has a
    letter column or a provider name column."""
    for i, row in enumerate(rows):
        vals = [norm(v) for v in row]
        if any(v in COL_ALIASES["letter"] for v in vals) or \
           any(v in COL_ALIASES["name"] for v in vals):
            return i, vals
    return None, None


def find_col(hdr, key):
    for alias in COL_ALIASES[key]:
        for j, h in enumerate(hdr):
            if h == alias:
                return j
    return None


def load_default_names(data_src):
    """letter -> name from DEFAULT_NAMES in data.js."""
    m = re.search(r"export const DEFAULT_NAMES = \{(.*?)\};", data_src, re.S)
    if not m:
        return {}
    return dict(re.findall(r'"([A-Z])":\s*"([^"]*)"', m.group(1)))


def name_key(s):
    return re.sub(r"[^a-z]", "", str(s).lower())


def num(v, default=0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def parse_workbook(path, letter_by_name):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    weeks = []
    unmatched = {}
    for name in wb.sheetnames:
        if name.strip().lower().startswith("cumulative"):
            print(f"  - Skipping sheet '{name}': cumulative tab")
            continue
        ws = wb[name]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        hdr_idx, hdr = find_header(rows)
        if hdr_idx is None:
            print(f"  ! Skipping sheet '{name}': no LETTER or PROVIDER NAME header")
            continue

        cols = {k: find_col(hdr, k) for k in COL_ALIASES}
        missing = [k for k in ("hours", "pts", "pthr", "esi1", "esi2",
                               "esi3", "esi4", "esi5") if cols[k] is None]
        if missing:
            print(f"  ! Skipping sheet '{name}': missing columns {missing}")
            continue
        if cols["letter"] is None and cols["name"] is None:
            print(f"  ! Skipping sheet '{name}': no letter or name column")
            continue

        try:
            start, end = parse_sheet_dates(name)
        except ValueError as e:
            print(f"  ! Skipping sheet '{name}': {e}")
            continue

        docs = []
        for row in rows[hdr_idx + 1:]:
            letter = None
            if cols["letter"] is not None:
                lt = row[cols["letter"]]
                if isinstance(lt, str) and len(lt.strip()) == 1 and lt.strip().isalpha():
                    letter = lt.strip().upper()
            if letter is None and cols["name"] is not None:
                nm = row[cols["name"]]
                if isinstance(nm, str) and nm.strip():
                    letter = letter_by_name.get(name_key(nm))
                    if letter is None:
                        unmatched.setdefault(nm.strip(), []).append(name)
                        continue
            if letter is None:
                continue

            hours = int(num(row[cols["hours"]]))
            pts = int(num(row[cols["pts"]]))
            pthr = num(row[cols["pthr"]], default=None)
            if pts == 0 or not pthr:
                pthr = None
            else:
                pthr = round(pthr, 2)
            esi = [int(num(row[cols[f"esi{k}"]])) for k in range(1, 6)]
            docs.append({"letter": letter, "hours": hours, "pts": pts,
                         "pthr": pthr, "esi": esi})

        if not docs:
            print(f"  ! Skipping sheet '{name}': no physician rows")
            continue
        weeks.append({"start": start, "end": end, "docs": docs, "sheet": name})

    if unmatched:
        print()
        print("  !! NAMES NOT FOUND IN DEFAULT_NAMES (rows skipped, fix before committing):")
        for nm, sheets in unmatched.items():
            print(f"     {nm}  (sheets: {', '.join(sorted(set(sheets)))})")
        print()
    return weeks


def extract_names(path):
    """Collect letter -> provider name from any sheet with both columns."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    names = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        hdr_idx, hdr = find_header(rows)
        if hdr_idx is None:
            continue
        c_name, c_letter = find_col(hdr, "name"), find_col(hdr, "letter")
        if c_name is None or c_letter is None:
            continue
        for row in rows[hdr_idx + 1:]:
            nm, lt = row[c_name], row[c_letter]
            if (isinstance(nm, str) and nm.strip()
                    and isinstance(lt, str) and len(lt.strip()) == 1
                    and lt.strip().isalpha()):
                names[lt.strip().upper()] = nm.strip()
    return names


def sync_names(data_src, names):
    """Rewrite DEFAULT_NAMES in data.js source. Returns (new_src, changed)."""
    if not names:
        return data_src, False
    m = re.search(r"export const DEFAULT_NAMES = \{.*?\};", data_src, re.S)
    if not m:
        print("  ! DEFAULT_NAMES not found in data.js, names NOT synced")
        return data_src, False
    js = ("export const DEFAULT_NAMES = { "
          + ", ".join(f'"{k}": "{v}"' for k, v in sorted(names.items()))
          + " };")
    if m.group(0) == js:
        return data_src, False
    return data_src.replace(m.group(0), js, 1), True


def render_week_js(week):
    start, end = week["start"], week["end"]
    wid = week_id(start)
    label, drange = labels_for(start, end)
    docs = week["docs"]
    active = sorted([d for d in docs if d["pts"] > 0],
                    key=lambda d: (-d["pthr"], -d["pts"]))
    inactive = [d for d in docs if d["pts"] == 0]

    hours = ", ".join(f"{d['letter']}: {d['hours']}" for d in active + inactive)
    lines = []
    for d in active + inactive:
        pthr = "null" if d["pthr"] is None else f"{d['pthr']:.2f}"
        e = d["esi"]
        lines.append(
            f'      {{ letter: "{d["letter"]}", pts: {d["pts"]}, pthr: {pthr}, '
            f'esi1: {e[0]}, esi2: {e[1]}, esi3: {e[2]}, esi4: {e[3]}, esi5: {e[4]} }},'
        )
    body = "\n".join(lines)
    return (f'  {{\n'
            f'    id: "{wid}",\n'
            f'    label: "{label}",\n'
            f'    dateRange: "{drange}",\n'
            f'    hoursWorked: {{ {hours} }},\n'
            f'    physicians: [\n{body}\n    ]\n'
            f'  }},\n')


def find_anomalies(week):
    notes = []
    for d in week["docs"]:
        if d["pts"] > 0:
            if d["pthr"] is not None and d["pthr"] < MIN_PTHR_FLAG:
                notes.append(f"{d['letter']}: pt/hr {d['pthr']:.2f} "
                             f"({d['pts']} pts / {d['hours']} hrs) is unusually low")
            if 0 < d["hours"] < MIN_HOURS_FLAG:
                notes.append(f"{d['letter']}: only {d['hours']} hours worked")
        if d["pts"] > 0 and sum(d["esi"]) != d["pts"]:
            notes.append(f"{d['letter']}: ESI counts sum to {sum(d['esi'])} "
                         f"but pts = {d['pts']}")
        if d["pts"] == 0 and d["hours"] > 0:
            notes.append(f"{d['letter']}: {d['hours']} hours but 0 patients")
    return notes


def newest_spreadsheet():
    """Most recently modified .xlsx under Provider PPH data, one level up from the repo."""
    pph = os.path.join(os.path.dirname(REPO), "Provider PPH data")
    if not os.path.isdir(pph):
        sys.exit(f"Can't find spreadsheet folder: {pph}\n"
                 "Pass a file path explicitly instead.")
    candidates = []
    for root, _dirs, files in os.walk(pph):
        for f in files:
            if f.lower().endswith(".xlsx") and not f.startswith("~$"):
                p = os.path.join(root, f)
                candidates.append((os.path.getmtime(p), p))
    if not candidates:
        sys.exit(f"No .xlsx files found under {pph}")
    return max(candidates)[1]


def main():
    if len(sys.argv) > 2:
        print(__doc__)
        sys.exit(1)
    if len(sys.argv) == 2:
        xlsx = os.path.expanduser(sys.argv[1])
    else:
        xlsx = newest_spreadsheet()
        print(f"Using newest spreadsheet: {xlsx}\n")
    if not os.path.exists(xlsx):
        sys.exit(f"File not found: {xlsx}")
    if not os.path.exists(DATA_JS):
        sys.exit(f"Can't find {DATA_JS}, is this script in the repo root?")

    data_src = open(DATA_JS, encoding="utf-8").read()

    # Names from data.js, then overlay any letter/name pairs in the sheet,
    # so sheets without a letter column can still be resolved.
    default_names = load_default_names(data_src)
    sheet_names = extract_names(xlsx)
    merged = dict(default_names)
    merged.update(sheet_names)
    letter_by_name = {name_key(v): k for k, v in merged.items()}

    weeks = parse_workbook(xlsx, letter_by_name)
    new = [w for w in weeks if week_id(w["start"]) not in data_src]

    print(f"Spreadsheet contains {len(weeks)} week(s); "
          f"{len(new)} new, {len(weeks) - len(new)} already in data.js.")

    # --- Weeks ---
    if new:
        new.sort(key=lambda w: w["start"], reverse=True)
        blocks = "".join(render_week_js(w) for w in new)
        anchor = "export const SEED_WEEKS = ["
        if anchor not in data_src:
            sys.exit("SEED_WEEKS anchor not found in data.js, NO CHANGES made.")
        data_src = data_src.replace(anchor, anchor + "\n" + blocks, 1)

    # --- Names ---
    data_src, names_changed = sync_names(data_src, sheet_names)
    if sheet_names:
        status = "updated" if names_changed else "already current"
        print(f"Provider names found for {len(sheet_names)} letters ({status}).")
    else:
        print("No letter/name pairs in this spreadsheet; names untouched.")

    if not new and not names_changed:
        print("Nothing to do. data.js and App.jsx untouched.")
        return

    open(DATA_JS, "w", encoding="utf-8").write(data_src)

    # --- Version bump ---
    app_src = open(APP_JSX, encoding="utf-8").read()
    m = re.search(r"const DATA_VERSION = (\d+)", app_src)
    if not m:
        sys.exit("DATA_VERSION not found in App.jsx. data.js updated, "
                 "but you must bump the version manually!")
    old_v, new_v = int(m.group(1)), int(m.group(1)) + 1
    open(APP_JSX, "w", encoding="utf-8").write(
        app_src.replace(f"const DATA_VERSION = {old_v}",
                        f"const DATA_VERSION = {new_v}", 1))

    print()
    for w in new:
        label, _ = labels_for(w["start"], w["end"])
        active = [d for d in w["docs"] if d["pts"] > 0]
        top = max(active, key=lambda d: d["pthr"])
        total = sum(d["pts"] for d in active)
        print(f"ADDED {label}  ({week_id(w['start'])})  from sheet '{w['sheet']}'")
        print(f"  {len(active)} active physicians, {total} total patients")
        print(f"  Top producer: {top['letter']} ({merged.get(top['letter'], '?')}) "
              f"at {top['pthr']:.2f} pt/hr")
        notes = find_anomalies(w)
        if notes:
            print("  FLAGS to review (consider confirming with Wendy):")
            for n in notes:
                print(f"    - {n}")
        print()

    if names_changed:
        print("DEFAULT_NAMES synced from spreadsheet:")
        for k in sorted(sheet_names):
            print(f"  {k}: {sheet_names[k]}")
        print()

    print(f"App.jsx: DATA_VERSION {old_v} -> {new_v}")
    print()
    print("Files updated. Next steps (NOT done automatically):")
    print('  git add src/data.js src/App.jsx update_leaderboard.py')
    print(f'  git commit -m "weekly data update, DATA_VERSION {new_v}"')
    print("  git push  (then verify on Vercel before emailing the group)")


if __name__ == "__main__":
    main()
